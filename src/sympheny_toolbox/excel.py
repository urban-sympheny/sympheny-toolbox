"""Excel helpers for Sympheny input/output files, based on openpyxl.

These helpers replicate the sheet layouts used by the Sympheny platform:
plain record sheets (header row + data rows) and profile sheets
(title row, ``Time step`` header row, then the hourly values).
"""

from __future__ import annotations

import io
import warnings
from pathlib import Path
from typing import IO, Any

from openpyxl import load_workbook


ExcelSource = str | Path | bytes | IO[bytes]


def _load(source: ExcelSource) -> Any:
    buffer: str | Path | IO[bytes] = io.BytesIO(source) if isinstance(source, bytes) else source
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return load_workbook(buffer, read_only=True, data_only=True)


def sheet_names(source: ExcelSource) -> list[str]:
    """Return the sheet names of an Excel workbook."""
    workbook = _load(source)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_records(source: ExcelSource, sheets: list[str]) -> dict[str, list[dict[str, Any]]]:
    """Read record sheets (first row = header) into ``{sheet: [row dicts]}``."""
    workbook = _load(source)
    try:
        result: dict[str, list[dict[str, Any]]] = {}
        for sheet in sheets:
            rows = workbook[sheet].iter_rows(values_only=True)
            header = [str(cell) if cell is not None else "" for cell in next(rows, ())]
            result[sheet] = [dict(zip(header, row)) for row in rows if any(cell is not None for cell in row)]
        return result
    finally:
        workbook.close()


def read_profile_sheets(source: ExcelSource, sheets: list[str]) -> dict[str, dict[str, list[Any]]]:
    """Read profile sheets (header on the second row) into ``{sheet: {column: values}}``.

    The ``Time step`` index column is excluded, matching the layout of Sympheny
    result files.
    """
    workbook = _load(source)
    try:
        result: dict[str, dict[str, list[Any]]] = {}
        for sheet in sheets:
            rows = workbook[sheet].iter_rows(values_only=True)
            next(rows, None)  # title row above the header
            header = [str(cell) if cell is not None else "" for cell in next(rows, ())]
            columns: dict[str, list[Any]] = {name: [] for name in header if name != "Time step"}
            for row in rows:
                for name, cell in zip(header, row):
                    if name != "Time step":
                        columns[name].append(cell)
            result[sheet] = columns
        return result
    finally:
        workbook.close()
