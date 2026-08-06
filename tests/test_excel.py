"""Tests for the openpyxl-based Excel helpers."""

from __future__ import annotations

import io

from openpyxl import Workbook

from sympheny_toolbox import excel


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_sheet_names_lists_every_sheet() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    workbook.create_sheet("Mode 1")

    assert excel.sheet_names(_workbook_bytes(workbook)) == ["Data", "Mode 1"]


def test_read_records_skips_blank_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Data"
    sheet.append(["Name", "Value"])
    sheet.append(["a", 1])
    sheet.append([None, None])
    sheet.append(["b", None])

    records = excel.read_records(_workbook_bytes(workbook), ["Data"])["Data"]

    assert records == [{"Name": "a", "Value": 1}, {"Name": "b", "Value": None}]


def test_read_profile_sheets_excludes_time_step_column() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Mode 1"
    sheet.append(["Result profiles", None, None])  # title row
    sheet.append(["Time step", "Gas boiler", "PV"])
    sheet.append([1, 10.0, 0.0])
    sheet.append([2, 12.5, 3.5])

    result = excel.read_profile_sheets(_workbook_bytes(workbook), ["Mode 1"])

    assert result == {"Mode 1": {"Gas boiler": [10.0, 12.5], "PV": [0.0, 3.5]}}
